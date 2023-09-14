import tkinter as tk
from tkinter import messagebox
from openpyxl import load_workbook

def buscar_nome(event=None):
    nome = entrada_nome.get()

    # Verifique se o nome não está vazio
    if not nome:
        messagebox.showerror("Erro", "Digite um nome para buscar.")
        return

    try:
        # Abra o arquivo da planilha
        arquivo = 'C:\\Users\\RECEBIMENTO 1\\Documents\\codpai.xlsx'
        planilha = load_workbook(arquivo, read_only=True)

        nome_encontrado = None
        codigo = None

        # Percorra todas as abas/planilhas no arquivo
        for sheet_name in planilha.sheetnames:
            planilha_ativa = planilha[sheet_name]

            for linha in planilha_ativa.iter_rows(values_only=True):
                if nome.lower() in str(linha[1]).lower():
                    nome_encontrado = linha[1]
                    codigo = linha[0]
                    break

            if nome_encontrado:
                break

        # Exiba o resultado na interface
        if nome_encontrado:
            resultado.config(text=f"Nome: {nome_encontrado}\nCódigo: {codigo}")
        else:
            resultado.config(text="Nome não encontrado nas planilhas.")

    except Exception as e:
        messagebox.showerror("Erro", f"Ocorreu um erro ao abrir o arquivo da planilha: {e}")

# Crie a janela principal
janela = tk.Tk()
janela.title("Busca De Código Pai")

# Defina as dimensões da janela (largura x altura)
largura_janela = 400
altura_janela = 200
janela.geometry(f"{largura_janela}x{altura_janela}")

# Crie a entrada para o nome
entrada_nome = tk.Entry(janela, width=30)
entrada_nome.pack(pady=10)

# Vincule a função buscar_nome à tecla "Enter" na entrada de texto
entrada_nome.bind("<Return>", buscar_nome)

# Crie o botão de busca
botao_buscar = tk.Button(janela, text="Buscar", command=buscar_nome)
botao_buscar.pack()

# Crie um rótulo para exibir o resultado
resultado = tk.Label(janela, text="")
resultado.pack(pady=10)

janela.mainloop()
